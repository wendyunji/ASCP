import os
import torch
import collections
import torch.nn as nn
import torch.nn.functional as F
import torch.optim as optim
import numpy as np
import random
import openpyxl
from datetime import datetime
import csv
from embedData import readXlsx, embedFlightData
from utils import checkConnection, get_reward, update_state
from CrewPairingEnv import CrewPairingEnv

# Hyperparameters
learning_rate = 0.005
gamma = 0.98
buffer_limit = 50000
batch_size = 32
INF = 999999999999999


class ReplayBuffer():
    def __init__(self):
        self.buffer = collections.deque(maxlen=buffer_limit)
    
    def put(self, transition):
        self.buffer.append(transition)
    
    def sample(self, n):
        mini_batch = random.sample(self.buffer, n)
        s_lst, a_lst, r_lst, s_prime_lst, done_mask_lst = [], [], [], [], []
        
        for transition in mini_batch:
            s, a, r, s_prime, done_mask = transition
            s_lst.append(s)
            a_lst.append([a])
            r_lst.append([r])
            s_prime_lst.append(s_prime)
            done_mask_lst.append([done_mask])

        return torch.tensor(s_lst, dtype=torch.float), torch.tensor(a_lst), \
               torch.tensor(r_lst), torch.tensor(s_prime_lst, dtype=torch.float), \
               torch.tensor(done_mask_lst)
    
    def size(self):
        return len(self.buffer)

class Qnet(nn.Module):
    def __init__(self, NN_Size):
        super(Qnet, self).__init__()
        self.fc1 = nn.Linear(NN_Size, 128)
        self.fc2 = nn.Linear(128, 128)
        self.fc3 = nn.Linear(128, 2)

    def forward(self, x):
        x = torch.tensor(x, dtype=torch.float32)

        x = F.relu(self.fc1(x))
        x = F.relu(self.fc2(x))
        x = self.fc3(x)
        return x

            
def train(q, q_target, memory, optimizer):
    for i in range(10):
        s,a,r,s_prime,done_mask = memory.sample(batch_size)

        q_out = q(s)
        q_a = q_out.gather(1,a)
        max_q_prime = q_target(s_prime).max(1)[0].unsqueeze(1)
        target = r + gamma * max_q_prime * done_mask
        loss = F.smooth_l1_loss(q_a, target)
        
        optimizer.zero_grad()
        loss.backward()
        optimizer.step()

def print_xlsx(output, output_pairing_filename):
    workbook = openpyxl.Workbook()
    sheet = workbook.active

    # Pairing data 제목 추가
    sheet.cell(row=1, column=1, value="Pairing data")

    # 데이터를 엑셀에 쓰기
    for row_index, row_data in enumerate(output, start=2):  # 첫 번째 행은 이미 Pairing data로 사용되었으므로 2부터 시작
        # 각 행의 첫 열에는 1부터 시작하는 인덱스 추가
        sheet.cell(row=row_index, column=1, value=row_index - 1)

        # 나머지 데이터 추가
        for col_index, value in enumerate(row_data, start=2):  # 각 행의 두 번째 열부터 시작
            sheet.cell(row=row_index, column=col_index, value=value)

    workbook.save(output_pairing_filename)

def main():
    # 사용자로부터 실행할 데이터의 년월일, 에피소드 수, 실행 ID 입력 받기
    month = input("Enter the month of the input file (e.g., 201406): ")
    episodes = int(input("Enter the number of episodes to run: "))
    excutionId = input("Enter the excution ID (eg., 20240704-1): ")

    current_directory = os.path.dirname(__file__)
    
    # 요구 디렉토리
    eval_directory = os.path.join(current_directory, '../eval')
    models_directory = os.path.join(current_directory, '../models')
    # 디렉토리가 없으면 만들기
    if not os.path.exists(eval_directory):
        os.makedirs(eval_directory)
    
    # 데이터 입력 받기
    path = os.path.abspath(os.path.join(current_directory, '../input'))
    readXlsx(path, f'/ASCP_Data_Input_{month}.xlsx')

    # 데이터 임베딩
    flight_list, V_f_list, NN_size, airport_total = embedFlightData(path)
    print('Data Imported')
    print("Number of Flights :", len(flight_list))

    # Load Crew Pairing Environment
    env = CrewPairingEnv(V_f_list, flight_list, airport_total)
    q = Qnet(NN_size)
    
    # eval에 사용할 저장된 모델 불러오기
    loaded_model = torch.load(os.path.join(models_directory, f'dqn_model_{month}_{episodes}_{excutionId}.pth'))
    q.load_state_dict(loaded_model)
    q.eval()

    time = datetime.now()
    score = 0
    output = []

    s, _ = env.reset()  # V_p 출발공항, V_f 도착공항
    done = False
    
    while not done:
        a = q.forward(s).argmax().item()
        # a = random.randint(0,1) # rule-based 코드

        s_prime, r, done, truncated, info, output = env.step(action=a)

        s = s_prime     # action에 의해 바뀐 flight
        score += r
    
    print(f"score : {score:.2f}")

    env.close()
    
    # 훈련된 모델로 생성된 eval 페어링을 xlsx로 eval 디렉토리에 저장
    print_xlsx(output, os.path.join(eval_directory, f'eval_pairing_{month}_{episodes}_{excutionId}.xlsx'))
    
if __name__ == '__main__':
    main()